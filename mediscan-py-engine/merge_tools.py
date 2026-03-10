from datetime import datetime
from copy import copy, deepcopy

import sys
import os
import openpyxl


def merge_files_into_main_worksheet(files):

    # 1- Initiate main workbook and worksheet + settings
    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
        
    merged_ws.title = "Merged_Patients"
    merged_ws.sheet_view.rightToLeft = True # right to left sheet

    current_row_offset = 1 
    headers_set = False
    
    
    # 2- Merge all files
    for file in files:
            
        # In the web version, the "file" was an object of file from fastapi
        # but since file here is just a string of the file_path 
        # it can just be passed normaly to openpyxl
            
        source_wb = openpyxl.load_workbook(filename=file)
        source_ws = source_wb.active
            
        # 2- copies the headers only from the first file
        if not headers_set:
            for col_index, cell in enumerate(source_ws[1], 1):
                new_cell = merged_ws.cell(row=1, column=col_index, value=cell.value)
                
            # extra from me: copy basic styling
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                    
            # 2.5- adjust the columns' width
            for col_letter, col_dim in source_ws.column_dimensions.items():
                merged_ws.column_dimensions[col_letter].width = col_dim.width
                
            headers_set = True
            
            
        # 3- copy data rows            
        rows = list(source_ws.rows)
        num_rows = len(rows)
            
        if num_rows > 1:
            for row_index, row in enumerate(rows[1:], 1):
                target_row_index = current_row_offset + row_index
                
                # copies cell values
                for col_index, cell in enumerate(row, 1):
                    merged_ws.cell(row=target_row_index, column=col_index, value=cell.value)
                    
                merged_ws.row_dimensions[target_row_index].height = source_ws.row_dimensions[row_index + 1].height
                
        # 4- copying images (find images and move them to the new row)
        if hasattr(source_ws, '_images'):
            for image in source_ws._images:
                anchor_row = image.anchor._from.row + 1
                
                if anchor_row > 1:
                    # calc new position
                    new_anchor_row = current_row_offset + (anchor_row - 1)
                    new_img = deepcopy(image)
                        
                    # anchor of the CLONE
                    new_img.anchor._from.row = new_anchor_row - 1
                        
                    # image height
                    height_in_rows = image.anchor.to.row - image.anchor._from.row
                    new_img.anchor.to.row = new_anchor_row - 1 + height_in_rows
                        
                    merged_ws.add_image(new_img)
                
        # update the offset for the next file
        current_row_offset += (num_rows - 1)
        
    # 5- Save merged_wb directly to hard drive
    docs_folder = os.path.join(os.path.expanduser("~"), "Documents", "MediScan_Merged_Reports")
    os.makedirs(docs_folder, exist_ok=True) # creates the folder if it doesn't exist
        
    filename_str = datetime.now().strftime("Merged_Report_%Y-%m-%d_%H-%M-%S.xlsx")
    file_path = os.path.join(docs_folder, filename_str)

    merged_wb.save(file_path)

    return file_path
    



if __name__ == '__main__':
    try:
        
        # 1- Grab all file paths from Rust
        file_paths = sys.argv[1:]
        if not file_paths:
            raise Exception("No file paths were provided to the engine.")
        
        if len(file_paths) < 2:
            raise Exception("You need to provide at least 2 files to merge.")


        valid_files = []
        # 2- Check if files are valid
        for file in file_paths:
            if not os.path.exists(file):
                continue
            
            # "file" here is a string, not an object as in the web with fastapi
            if not file.lower().endswith(('.xlsx', '.xls')):
                continue
            
            valid_files.append(file)

        # to catch an edge case 
        if len(valid_files) < 2:
            raise Exception("You need to provide at least 2 valid Excel files to merge.")

        # 3-Process files and capture the saved path
        saved_path = merge_files_into_main_worksheet(valid_files)
        
        # 4- Message to listen to for Rust
        print(f"===MEDISCAN_SUCCESS==={saved_path}===END===")
    

    except Exception as e:
        print(f"===MEDISCAN_ERROR==={str(e)}===END===")
        sys.exit(1)        
